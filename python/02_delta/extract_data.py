import openpyxl, json

wb = openpyxl.load_workbook('Hedge.xlsm', data_only=True, read_only=True)

def cell(sheet, addr):
    return wb[sheet][addr].value

def label_lookup(sheet, max_row, max_col, label_col_offset=0):
    """Scan a sheet; return dict {label_string_stripped: [values_in_following_cols]}."""
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip():
                key = v.strip()
                rest = [x for x in row[i+1:] if x is not None]
                if rest and key not in out:
                    out[key] = rest
    return out

d = {}

# ---- Encoding ----
enc = label_lookup('Encoding', 62, 4)
d['WTI_spot']   = enc['WTI_spot'][0]
d['KRW_spot']   = enc['KRW_spot'][0]
d['r_US']       = enc['US_rate'][0]
d['r_KRW']      = enc['K_rate'][0]
d['WTI_vol_hist'] = enc['WTI_vol'][0]
d['KRW_vol']    = enc['KRW_vol'][0]
d['Monthly_Oil_Need'] = enc['Monthly_Oil_Need'][0]
d['Monthly_USD_Need'] = enc['Monthly_USD_Need'][0]
d['WACC']       = enc['WACC'][0]
d['Max_Budget'] = enc['Max_Budget'][0]
d['Maturity_Oil'] = enc['Maturity_Oil'][0]
d['Maturity_FX']  = enc['Maturity_FX'][0]
d['Stress_WTI'] = enc['Stress_WTI'][0]
d['Stress_KRW'] = enc['Stress_KRW'][0]

# ---- LSMC ----
lsmc = label_lookup('LSMC', 20, 15)
d['lambda_jump'] = cell('LSMC','B1')
d['jump_mean']   = cell('LSMC','B2')
d['jump_vol']    = cell('LSMC','B3')
d['mu1_P']       = cell('LSMC','B4')
d['mu2_P']       = cell('LSMC','B5')
d['KO_upper']    = cell('LSMC','B6')
d['KO_lower']    = cell('LSMC','B7')
d['S1_0']        = cell('LSMC','B8')
d['S2_0']        = cell('LSMC','B9')
d['diff_vol_WTI'] = cell('LSMC','B10')
d['vol_KRW']     = cell('LSMC','B11')
d['corr']        = cell('LSMC','B12')
d['n_steps']     = cell('LSMC','B13')
d['T_total']     = cell('LSMC','B14')
d['base_premium']   = lsmc['Base Premium'][0]
d['stress_premium'] = lsmc['Stress Premium'][0]
d['WTI_delta_krw']  = lsmc['WTI Delta (KRW)'][0]
d['FX_delta_krw']   = lsmc['FX Delta (KRW)'][0]
d['WTI_share']      = lsmc['WTI Share (Ratio)'][0]
d['FX_share']       = lsmc['FX Share (Ratio)'][0]
d['WTI_base_premium_shapley'] = lsmc['WTI Base Premium (Shapley)'][0]
d['FX_base_premium_shapley']  = lsmc['FX Base Premium (Shapley)'][0]

# ---- American / European / Asian summary panels (label-based) ----
am = label_lookup('American_Delta', 45, 20)
d['american'] = {
    'sim_runs': cell('American_Delta','B1'),
    'base_premium': am['Base Premium (KRW):'][0],
    'total_premium': am['Total Premium (KRW):'][0],
    'wti_delta_ratio': am['WTI Delta Ratio'][0],
    'fx_delta_ratio': am['FX Delta Ratio'][0],
    'wti_premium': am['WTI Premium (LSMC)'][0],
    'fx_premium': am['FX Premium (LSMC)'][0],
    'steps': am['Steps (N=260wks)'][0],
    'opt_mean': am['Mean'][0], 'tot_mean': am['Mean'][1],
    'opt_std':  am['Std Dev'][0], 'tot_std':  am['Std Dev'][1],
    'opt_min':  am['Minimum'][0], 'tot_min':  am['Minimum'][1],
    'opt_max':  am['Maximum'][0], 'tot_max':  am['Maximum'][1],
    'opt_skew': am['Skewness'][0], 'tot_skew': am['Skewness'][1],
    'opt_kurt': am['Kurtosis'][0], 'tot_kurt': am['Kurtosis'][1],
    'opt_median': am['Median'][0],'tot_median': am['Median'][1],
    'avg_jump_loss': am['Avg Jump Loss'][0],
    'ko_rate': am['KO Rate'][0],
    'exercise_rate': am['Early Exercise Rate'][0],
}
eu = label_lookup('European_Delta', 40, 20)
d['european'] = {
    'sim_runs': cell('European_Delta','B1'),
    'strike_wti': eu['Strike (K_WTI):'][0],
    'base_premium': eu['Base Premium (KRW):'][0],
    'total_premium': eu['Total Premium (KRW):'][0],
    'strike_fx': eu['KRW Strike'][0],
    'opt_mean': eu['Mean'][0], 'tot_mean': eu['Mean'][1],
    'opt_std':  eu['Std Dev'][0], 'tot_std':  eu['Std Dev'][1],
    'ko_rate':  eu['KO Rate'][0],
}
asx = label_lookup('Asian_Delta', 40, 21)
d['asian'] = {
    'sim_runs': cell('Asian_Delta','B1'),
    'strike_wti': asx['Strike (K_WTI):'][0],
    'base_premium': asx['Base Premium (KRW):'][0],
    'total_premium': asx['Total Premium (KRW):'][0],
    'maturity_wti': asx['Maturity WTI (yr)'][0],
    'maturity_fx': asx['Maturity FX (yr)'][0],
    'wti_steps': asx['WTI Steps (N)'][0],
    'fx_steps': asx['FX Steps (N)'][0],
    'mc_paths': asx['MC Paths'][0],
    'opt_mean': asx['Mean'][0], 'tot_mean': asx['Mean'][1],
    'opt_std':  asx['Std Dev'][0], 'tot_std':  asx['Std Dev'][1],
    'ko_rate':  asx['KO Rate'][0],
}

# ---- PayoffParityCheck (label-based) ----
pp = label_lookup('PayoffParityCheck', 26, 7)
d['parity'] = {
    'n': cell('PayoffParityCheck','B2'),
    'full': {
        'am_mean': pp['Mean (KRW)'][0], 'eu_mean': pp['Mean (KRW)'][1], 'as_mean': pp['Mean (KRW)'][2],
        'eu_minus_am': pp['Mean (KRW)'][3], 'as_minus_am': pp['Mean (KRW)'][4],
        'am_std': pp['StdDev (KRW)'][0], 'eu_std': pp['StdDev (KRW)'][1], 'as_std': pp['StdDev (KRW)'][2],
    },
}
# clean subset appears as a second "Mean (KRW)"/"StdDev (KRW)" occurrence -- grab via raw rows instead
ws = wb['PayoffParityCheck']
raw = [row for row in ws.iter_rows(min_row=1, max_row=26, max_col=7, values_only=True)]
# row indices (1-based) per manual inspection
d['parity']['clean'] = {'n_clean': 3217,
    'am_mean': raw[11][1], 'eu_mean': raw[11][2], 'eu_minus_am': raw[11][4],
    'am_std': raw[12][1], 'eu_std': raw[12][2]}
d['parity']['bias_am_eu'] = raw[16][2]
d['parity']['bias_am_as'] = raw[17][2]

# ---- Girsanov ----
gs = label_lookup('Girsanov', 6, 3)
d['girsanov'] = {
    'reweighted_mean': gs['Reweighted (Q-paths x Girsanov LR)'][0], 'reweighted_std': gs['Reweighted (Q-paths x Girsanov LR)'][1],
    'direct_mean': gs['Direct (P-measure simulation)'][0], 'direct_std': gs['Direct (P-measure simulation)'][1],
    'residual_mean': gs['Residual (Mean diff)'][0],
    'residual_pct': gs['Residual (% of Direct Mean)'][0],
}

def rows(sheet, min_row, max_row, max_col):
    ws = wb[sheet]
    out = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True):
        if any(v is not None for v in row):
            out.append(list(row))
    return out

d['driftbias'] = rows('DriftBias', 2, 7, 6)
d['sharpesweep'] = rows('SharpeSweep', 2, 6, 4)
d['lambdarobustness'] = rows('LambdaRobustness', 2, 4, 12)
d['deltacheck'] = rows('DeltaCheck', 2, 4, 6)
d['deltafxsweep'] = rows('DeltaFXSweep', 2, 6, 7)
d['jumppremiumsens'] = rows('JumpPremiumSens', 2, 5, 4)

# ---- ITM Diagnostics summary per maturity ----
ws = wb['Diagnostics']
itm = {}
for r in ws.iter_rows(min_row=2, max_row=2731, max_col=4, values_only=True):
    if r[0] is None: continue
    mat, step, tot, p = r
    itm.setdefault(mat, []).append((tot, p))
d['itm_diag'] = {str(k): v for k, v in itm.items()}

# ---- Raw_Timeseries (historical WTI / KRW) for the motivating figure ----
ws = wb['Raw_Timeseries']
ts = {'dates': [], 'wti': [], 'krw': []}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3, values_only=True):
    if row[0] is None or row[1] is None:
        continue
    ts['dates'].append(row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0]))
    ts['wti'].append(row[1])
    ts['krw'].append(row[2])
d['raw_timeseries'] = ts

with open('scraped_data.json', 'w', encoding='utf-8') as f:
    json.dump(d, f, indent=2, default=str)

print("OK. Spot check:")
print("base_premium", d['base_premium'])
print("american:", {k:v for k,v in d['american'].items() if k in ('opt_mean','tot_mean','ko_rate','exercise_rate','avg_jump_loss')})
print("european:", d['european'])
print("asian:", d['asian'])
print("parity full:", d['parity']['full'])
print("parity clean:", d['parity']['clean'])
print("parity bias:", d['parity']['bias_am_eu'], d['parity']['bias_am_as'])
print("girsanov:", d['girsanov'])
print("deltacheck:", d['deltacheck'])
print("deltafxsweep:", d['deltafxsweep'])
