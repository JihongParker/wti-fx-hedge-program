import json
import openpyxl

WB = "Hedge 복사본.xlsm"

wb = openpyxl.load_workbook(WB, data_only=True, read_only=True)

def cell(sheet, addr):
    return wb[sheet][addr].value

def label_lookup(sheet, max_row, max_col):
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip():
                key = v.strip()
                rest = [x for x in row[i+1:] if x is not None]
                if rest and key not in out:
                    out[key] = rest
    return out

def series(sheet, max_col):
    ws = wb[sheet]
    header_row = None
    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True), start=1):
        if row[0] == 'Step':
            header_row = row
            continue
        if header_row is None:
            continue
        if row[0] is None:
            continue
        rows.append(row)
    return list(header_row), rows

d = {}

# ---- version / integrity check (soft — report, don't hard-block) ----
spec_row = cell('CFH_Comparison', 'A1')
d['spec_version'] = spec_row
required_sheets = ['CFH_Comparison', 'CFH_Economic', 'CFH_A_Ledger', 'CFH_B_Ledger_WTI', 'CFH_B_Ledger_FX', 'CFH_SFP_Proforma', 'CFH_Inputs']
missing = [s for s in required_sheets if s not in wb.sheetnames]
if missing:
    raise RuntimeError(f"Missing required CFH sheets: {missing}. Re-run Run_CFH_Accounting_Engine first.")

# ---- shared market parameters (Encoding / LSMC — same as Delta_Simulation) ----
wsE = wb['Encoding']
d['S1_0'] = cell('Encoding', 'B2')
d['S2_0'] = cell('Encoding', 'B3')
d['r_US'] = cell('Encoding', 'B4')
d['r_KRW'] = cell('Encoding', 'B5')
d['Monthly_Oil_Need'] = cell('Encoding', 'B9')
d['Monthly_USD_Need'] = cell('Encoding', 'B10')
d['T_WTI'] = cell('Encoding', 'B15')
d['T_FX'] = cell('Encoding', 'B16')

wsL = label_lookup('LSMC', 20, 15)
d['lambda_jump'] = cell('LSMC', 'B1')
d['jump_mean'] = cell('LSMC', 'B2')
d['jump_vol'] = cell('LSMC', 'B3')
d['KO_upper'] = cell('LSMC', 'B6')
d['KO_lower'] = cell('LSMC', 'B7')
d['diff_vol_WTI'] = cell('LSMC', 'B10')
d['vol_KRW'] = cell('LSMC', 'B11')
d['corr'] = cell('LSMC', 'B12')

# ---- CFH_Inputs ----
ci = label_lookup('CFH_Inputs', 10, 3)
d['n_acct'] = ci['n_acct (realised paths)'][0]
d['n_sens'] = ci['n_sens (FV-validation reprice)'][0]

# ---- CFH_Comparison (H1-H4 + FV validation) ----
ws = wb['CFH_Comparison']
raw = [list(row) for row in ws.iter_rows(min_row=1, max_row=32, max_col=7, values_only=True)]
comp = label_lookup('CFH_Comparison', 32, 6)
d['comparison'] = {
    'spec_header': raw[0][0],
    'premium_outflow_check': comp['Premium Outflow Check'][0],
    'A_FV0_mean': comp['A_FV(0) Mean'][0],
    'H1_A_mean_abs_ineff': comp['Structure A Mean |Cum Ineff|'][0],
    'H1_B_mean_abs_ineff': comp['Structure B Mean |Cum Ineff|'][0],
    'H2_sigma_econ_A': comp['Structure A Full Horizon sigma_econ'][0],
    'H2_sigma_econ_B': comp['Structure B Full Horizon sigma_econ'][0],
    'H3_KO_probability': comp['Knock-Out Probability'][0],
    'H3_VaR99_naked_A': comp['Structure A Naked Exposure VaR99'][0],
    'H3_VaR99_naked_B': comp['Structure B Naked Exposure VaR99'][0],
    'H3_A_OCI_to_PL_reclass': comp['A Structure OCI->P&L Reclassification Amount'][0],
    'H4_n_deriv_lines_A': comp['Number of derivative lines (A vs B)'][0],
    'H4_n_deriv_lines_B': comp['Number of derivative lines (A vs B)'][1],
    'H4_B_mean_postKO_FVTPL': comp['B mean post-KO FVTPL P&L (KO paths)'][0],
    'H4_B_std_postKO_FVTPL': comp['B std post-KO FVTPL P&L (KO paths)'][0],
}

# FV validation matrix rows (tau/T=0, 0.5, 0.9)
fv_rows = []
for row in raw:
    if isinstance(row[0], (int, float)) and row[0] in (0, 0.1, 0.5, 0.9) and isinstance(row[1], (int, float)):
        fv_rows.append(row)
d['fv_validation'] = fv_rows  # [tau/T, A_Beta_FV, A_Reprice_FV, B1_Beta_FV, B1_Reprice_FV, A_Pct_Diff]

# ---- CFH_Economic (full 217-step time series) ----
hdr, rows = series('CFH_Economic', 6)
d['economic_series'] = {'header': hdr, 'rows': rows}

# ---- CFH_A_Ledger ----
hdr, rows = series('CFH_A_Ledger', 12)
d['a_ledger'] = {'header': hdr, 'rows': rows}

# ---- CFH_B_Ledger_WTI ----
hdr, rows = series('CFH_B_Ledger_WTI', 6)
d['b1_ledger'] = {'header': hdr, 'rows': rows}

# ---- CFH_B_Ledger_FX ----
hdr, rows = series('CFH_B_Ledger_FX', 8)
d['b2_ledger'] = {'header': hdr, 'rows': rows}

# ---- CFH_SFP_Proforma ----
hdr, rows = series('CFH_SFP_Proforma', 11)
d['sfp'] = {'header': hdr, 'rows': rows}

with open('cfh_data.json', 'w', encoding='utf-8') as f:
    json.dump(d, f, indent=2, default=str)

print("Wrote cfh_data.json")
print("spec header:", d['comparison']['spec_header'])
print("H1 A/B:", d['comparison']['H1_A_mean_abs_ineff'], d['comparison']['H1_B_mean_abs_ineff'])
print("H2 A/B:", d['comparison']['H2_sigma_econ_A'], d['comparison']['H2_sigma_econ_B'])
print("H3 KO/VaRA/VaRB/Reclass:", d['comparison']['H3_KO_probability'], d['comparison']['H3_VaR99_naked_A'], d['comparison']['H3_VaR99_naked_B'], d['comparison']['H3_A_OCI_to_PL_reclass'])
print("economic_series rows:", len(d['economic_series']['rows']))
print("a_ledger rows:", len(d['a_ledger']['rows']))
print("b1_ledger rows:", len(d['b1_ledger']['rows']))
print("b2_ledger rows:", len(d['b2_ledger']['rows']))
print("sfp rows:", len(d['sfp']['rows']))
print("T_WTI/T_FX:", d['T_WTI'], d['T_FX'])
